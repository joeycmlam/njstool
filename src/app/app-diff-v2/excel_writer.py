import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os


class ExcelWriter:
    """
    A class to write comparison results to an Excel file.
    """
    def __init__(self, logger, output_config=None):
        self.logger = logger
        self.output_config = output_config or {}
        self.output_path = self.output_config.get("path", "./result")
        self.output_file_name = self.output_config.get("file", "comparison_results.xlsx")
        self.headers = self.output_config.get("headers", [
            "Row Number (File A)", "Row Number (File B)", "Hash Key",
            "Human-Readable Key", "Column", "File A Value", "File B Value", "Status"
        ])
        self.sheet_name = self.output_config.get("sheet_name", "Comparison Results")
        self.group_by = self.output_config.get("group_by", None)
        self.timestamped = self.output_config.get("timestamped", True)
        self.workbook = None
        
        # Ensure output directory exists
        os.makedirs(self.output_path, exist_ok=True)

    def _get_output_file_path(self):
        """
        Generate the complete output file path with timestamp if configured.
        """
        if self.timestamped:
            base_name = os.path.splitext(self.output_file_name)[0]
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"{base_name}_{timestamp}.xlsx"
        else:
            file_name = self.output_file_name
        return os.path.join(self.output_path, file_name)

    def write_source_context(self, source_context, worksheet_name):
        """
        Write source context to a worksheet.
        
        Args:
            source_context (list): List of dictionaries containing source context data
            worksheet_name (str): Name of the worksheet to create
        """
        if not self.workbook:
            self.workbook = openpyxl.Workbook()

        try:
            # Create new worksheet
            if worksheet_name in self.workbook.sheetnames:
                sheet = self.workbook[worksheet_name]
            else:
                sheet = self.workbook.create_sheet(title=worksheet_name)

            # Write headers if source_context is not empty
            if source_context and len(source_context) > 0:
                headers = ["Row Number"] + list(source_context[0].keys())
                for col_num, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    sheet.column_dimensions[cell.column_letter].width = 30
                    cell.alignment = Alignment(horizontal="center")

                # Write data
                for row_num, context in enumerate(source_context, start=2):
                    # Write row number
                    sheet.cell(row=row_num, column=1, value=row_num-1)
                    
                    # Write context values
                    for col_num, value in enumerate(context.values(), start=2):
                        sheet.cell(row=row_num, column=col_num, value=value)

            self.logger.info(f"Source context written to worksheet: {worksheet_name}")
        except Exception as e:
            self.logger.error(f"Error writing source context to worksheet {worksheet_name}: {e}")
            raise

    def write_to_excel(self, results):
        try:
            if not results:
                self.logger.warning("No results to write to Excel.")
                return

            # Validate results format
            required_fields = {"row_number_a", "row_number_b", "key", "composite_key", "column", 
                             "file_a_value", "file_b_value", "status"}
            if not all(field in results[0] for field in required_fields):
                self.logger.error("Results data is missing required fields.")
                raise ValueError("Invalid results data format.")

            output_file = self._get_output_file_path()
            if not self.workbook:
                self.workbook = openpyxl.Workbook()

            # Write comparison results
            if self.group_by:
                grouped_results = self._group_results(results, self.group_by)
                for group, group_results in grouped_results.items():
                    sheet = self.workbook.create_sheet(title=group)
                    self._write_headers(sheet, self.headers)
                    self._write_data(sheet, group_results)
            else:
                sheet = self.workbook.active
                sheet.title = self.sheet_name
                self._write_headers(sheet, self.headers)
                self._write_data(sheet, results)

            self.workbook.save(output_file)
            self.logger.info(f"Results written to {output_file}")
        except Exception as e:
            self.logger.exception(f"Error writing to Excel file: {e}")
            raise

    @staticmethod
    def _write_headers(sheet, headers):
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            sheet.column_dimensions[cell.column_letter].width = 30  # Increased width for better readability
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _write_data(sheet, results):
        for row_num, result in enumerate(results, start=2):
            sheet.cell(row=row_num, column=1, value=result["row_number_a"])
            sheet.cell(row=row_num, column=2, value=result["row_number_b"])
            sheet.cell(row=row_num, column=3, value=result["key"])
            sheet.cell(row=row_num, column=4, value=result["composite_key"])
            sheet.cell(row=row_num, column=5, value=result["column"])
            sheet.cell(row=row_num, column=6, value=result["file_a_value"])
            sheet.cell(row=row_num, column=7, value=result["file_b_value"])
            sheet.cell(row=row_num, column=8, value=result["status"])

    @staticmethod
    def _group_results(results, group_by):
        grouped = {}
        for result in results:
            key = result.get(group_by, "Other")
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(result)
        return grouped

    def _write_source_info(self, workbook):
        """
        Write source file information to a separate worksheet.
        """
        source_sheet = workbook.create_sheet(title="Source Files")
        
        # Set headers
        headers = ["File Type", "File Path", "Last Modified", "Size (bytes)"]
        for col_num, header in enumerate(headers, start=1):
            cell = source_sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            source_sheet.column_dimensions[cell.column_letter].width = 50
            cell.alignment = Alignment(horizontal="center")

        # Write file A info
        file_a_path = self.output_config.get("file_a_path", "N/A")
        file_a_stats = os.stat(file_a_path) if os.path.exists(file_a_path) else None
        source_sheet.cell(row=2, column=1, value="File A")
        source_sheet.cell(row=2, column=2, value=file_a_path)
        source_sheet.cell(row=2, column=3, value=datetime.fromtimestamp(file_a_stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S') if file_a_stats else "N/A")
        source_sheet.cell(row=2, column=4, value=file_a_stats.st_size if file_a_stats else "N/A")

        # Write file B info
        file_b_path = self.output_config.get("file_b_path", "N/A")
        file_b_stats = os.stat(file_b_path) if os.path.exists(file_b_path) else None
        source_sheet.cell(row=3, column=1, value="File B")
        source_sheet.cell(row=3, column=2, value=file_b_path)
        source_sheet.cell(row=3, column=3, value=datetime.fromtimestamp(file_b_stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S') if file_b_stats else "N/A")
        source_sheet.cell(row=3, column=4, value=file_b_stats.st_size if file_b_stats else "N/A")

    def _write_source_contents(self, workbook, sheet_title, file_path, columns):
        """
        Write the contents of a source file to a worksheet.
        """
        if not file_path or not os.path.exists(file_path):
            self.logger.warning(f"Source file not found: {file_path}")
            return

        try:
            sheet = workbook.create_sheet(title=sheet_title)

            # Create headers based on column definitions
            headers = ["Row Number"] + list(columns.keys())
            for col_num, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                sheet.column_dimensions[cell.column_letter].width = 20
                cell.alignment = Alignment(horizontal="center")

            # Read and write file contents
            with open(file_path, 'r', encoding='utf-8') as file:
                for row_num, line in enumerate(file, start=2):
                    # Write row number
                    sheet.cell(row=row_num, column=1, value=row_num-1)
                    
                    # Write column values
                    for col_num, (col_name, col_props) in enumerate(columns.items(), start=2):
                        start = col_props.get("start", 0)
                        length = col_props.get("length", 0)
                        value = line[start:start + length].strip() if start + length <= len(line) else ""
                        sheet.cell(row=row_num, column=col_num, value=value)

        except Exception as e:
            self.logger.error(f"Error writing source contents for {file_path}: {e}")
            raise
