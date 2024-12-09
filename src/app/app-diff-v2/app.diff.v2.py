import json
import openpyxl
from openpyxl.styles import Font
import logging
from datetime import datetime
import os


class Logger:
    """
    A class to configure and manage logging.
    """
    def __init__(self, log_dir="./log"):
        self.logger = self._initialize_logger(log_dir)

    @staticmethod
    def _initialize_logger(log_dir):
        os.makedirs(log_dir, exist_ok=True)
        log_file_name = f"{log_dir}/file_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        logger = logging.getLogger("FileComparison")
        logger.setLevel(logging.INFO)

        # File handler
        file_handler = logging.FileHandler(log_file_name)
        file_handler.setFormatter(logging.Formatter(Logger._log_format()))
        logger.addHandler(file_handler)

        # Console handler
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(logging.Formatter(Logger._log_format()))
        logger.addHandler(console_handler)

        return logger

    @staticmethod
    def _log_format():
        return "%(asctime)s - %(name)s - %(levelname)s - [%(funcName)s:%(lineno)d] - %(message)s"

    def get_logger(self):
        return self.logger


class Config:
    """
    A class to manage the configuration file.
    """
    def __init__(self, config_path):
        self.config_data = self._load_config(config_path)

    def _load_config(self, config_path):
        try:
            with open(config_path, 'r') as config_file:
                return json.load(config_file)
        except FileNotFoundError:
            raise FileNotFoundError(f"Config file not found: {config_path}")
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON format in config file: {e}")

    def get(self, key, default=None):
        return self.config_data.get(key, default)


class FileParser:
    """
    A class to parse and process input files.
    """
    def __init__(self, logger):
        self.logger = logger

    def parse_file(self, file_path, columns, skip_keys=None):
        """
        Parse a file into a list of records and a dictionary keyed by the composite key.

        :param file_path: Path to the file to parse.
        :param columns: Dictionary defining column positions, lengths, and flags.
        :param skip_keys: A set of composite keys to skip during parsing.
        :return: A tuple (record_dict, records)
        """
        record_dict = {}
        records = []
        skip_keys = skip_keys or set()
        key_columns = self._get_key_columns(columns)

        try:
            with open(file_path, 'r') as file:
                for line_no, line in enumerate(file, start=1):
                    record = self._parse_line(line, columns, line_no)
                    composite_key = self._construct_composite_key(record, key_columns)
                    if composite_key in skip_keys:
                        continue

                    records.append(record)
                    record_dict[composite_key] = record
        except FileNotFoundError:
            self.logger.error(f"File not found: {file_path}")
            raise
        except Exception as e:
            self.logger.exception(f"Error processing file {file_path}: {e}")
            raise

        self.logger.info(f"Parsed file: {file_path} with {len(records)} records.")
        return record_dict, records

    @staticmethod
    def _get_key_columns(columns):
        return [col for col, props in columns.items() if props.get("key", False)]

    @staticmethod
    def _parse_line(line, columns, line_no):
        record = {"row_number": line_no}
        for column_name, props in columns.items():
            if not props.get("skip", False):
                start = props["start"]
                length = props["length"]
                record[column_name] = line[start:start + length].strip()
        return record

    @staticmethod
    def _construct_composite_key(record, key_columns):
        return "-".join(str(record.get(column, "")) for column in key_columns)


class RecordComparator:
    """
    A class to compare records from two files.
    """
    def __init__(self, logger):
        self.logger = logger

    def compare_records(self, file_a_data, file_b_data, columns):
        """
        Compare records from two files based on the specified columns.

        :param file_a_data: Parsed data from File A.
        :param file_b_data: Parsed data from File B.
        :param columns: Dictionary defining column properties.
        :return: A list of comparison results.
        """
        results = []
        all_keys = set(file_a_data.keys()).union(file_b_data.keys())
        self.logger.info(f"Starting comparison with {len(all_keys)} composite keys.")

        for key in all_keys:
            record_a = file_a_data.get(key, {})
            record_b = file_b_data.get(key, {})

            row_number_a = record_a.get("row_number", "N/A")
            row_number_b = record_b.get("row_number", "N/A")

            for column_name, props in columns.items():
                if not props.get("skip", False):
                    results.append(self._compare_column(record_a, record_b, key, row_number_a, row_number_b, column_name))

        self.logger.info(f"Comparison completed with {len(results)} results.")
        return results

    @staticmethod
    def _compare_column(record_a, record_b, key, row_number_a, row_number_b, column_name):
        value_a = record_a.get(column_name, None)
        value_b = record_b.get(column_name, None)

        if value_a == value_b:
            status = "match"
        elif value_a is None:
            status = "missing in File A"
        elif value_b is None:
            status = "missing in File B"
        else:
            status = "not-match"

        return {
            "row_number_a": row_number_a,
            "row_number_b": row_number_b,
            "key": key,
            "column": column_name,
            "file_a_value": value_a,
            "file_b_value": value_b,
            "status": status
        }


class ExcelWriter:
    """
    A class to write comparison results to an Excel file.
    """
    def __init__(self, logger):
        self.logger = logger

    def write_to_excel(self, results, output_file):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Comparison Results"

            headers = ["Row Number (File A)", "Row Number (File B)", "Composite Key", "Column", "File A Value", "File B Value", "Status"]
            self._write_headers(sheet, headers)
            self._write_data(sheet, results)

            workbook.save(output_file)
            self.logger.info(f"Results written to {output_file}")
        except Exception as e:
            self.logger.exception(f"Error writing to Excel file {output_file}: {e}")
            raise

    @staticmethod
    def _write_headers(sheet, headers):
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

    @staticmethod
    def _write_data(sheet, results):
        for row_num, result in enumerate(results, start=2):
            sheet.cell(row=row_num, column=1, value=result["row_number_a"])
            sheet.cell(row=row_num, column=2, value=result["row_number_b"])
            sheet.cell(row=row_num, column=3, value=result["key"])
            sheet.cell(row=row_num, column=4, value=result["column"])
            sheet.cell(row=row_num, column=5, value=result["file_a_value"])
            sheet.cell(row=row_num, column=6, value=result["file_b_value"])
            sheet.cell(row=row_num, column=7, value=result["status"])


class FileComparisonApp:
    """
    The main orchestrator for the file comparison application.
    """
    def __init__(self, config_path):
        self.logger = Logger().get_logger()
        self.config = Config(config_path)
        self.parser = FileParser(self.logger)
        self.comparator = RecordComparator(self.logger)
        self.writer = ExcelWriter(self.logger)

    def run(self):
        try:
            file_a_path = self.config.get("file_a")
            file_b_path = self.config.get("file_b")
            columns = self.config.get("columns")
            skip_keys = set(self.config.get("skip_keys", []))

            # Get output path and file name from config
            output_path = self.config.get("output_path", "./")
            output_file_name = self.config.get("output_file", "comparison_results.xlsx")
            output_file = self._generate_full_output_path(output_path, output_file_name)

            # Parse both files
            file_a_data, _ = self.parser.parse_file(file_a_path, columns, skip_keys)
            file_b_data, _ = self.parser.parse_file(file_b_path, columns, skip_keys)

            # Compare records
            results = self.comparator.compare_records(file_a_data, file_b_data, columns)

            # Write results to Excel
            self.writer.write_to_excel(results, output_file)
        except Exception as e:
            self.logger.error(f"An error occurred during file comparison: {e}")
            raise

    @staticmethod
    def _generate_full_output_path(output_path, output_file_name):
        os.makedirs(output_path, exist_ok=True)
        timestamped_file_name = f"{os.path.splitext(output_file_name)[0]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return os.path.join(output_path, timestamped_file_name)


if __name__ == "__main__":
    app = FileComparisonApp("config.json")
    app.run()